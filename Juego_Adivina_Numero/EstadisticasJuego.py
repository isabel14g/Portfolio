import openpyxl
import os
import matplotlib.pyplot as plt 

# Ruta archivo
rutaArchivo = os.path.join(os.path.dirname(__file__), 'EstadisticasJuego.xlsx')

# Esta función sirve para guardar los resultados de una partida en Excel
def GuardarResultado(nombre, modo, dificultad, intentos, resultado):
    #Cargamos el archivo y seleccionamos la hoja que esté activa
    excelDocument = openpyxl.load_workbook(rutaArchivo)
    hoja = excelDocument.active

    # Vamos a escribir una nueva fila con los datos de cada jugador y guardarlo
    hoja.append([nombre, modo, dificultad, intentos, resultado])
    excelDocument.save(rutaArchivo)

# Función que muestra las estadísitcas que están en el archivo excel
def MostrarEstadisticas():
    if os.path.exists(rutaArchivo):
        excelDocument = openpyxl.load_workbook(rutaArchivo)
        hoja = excelDocument.active
        print("\n--- Estadísticas Totales del Juego ---")
        for fila in hoja.iter_rows(values_only=True):   # A continuación, recorremos las filas y devolvemos el valor real de la celda
            for celda in fila:
                print(celda, end=" | ")  
            print()
    else:
        print("\nNo hay estadísticas disponibles.")

# Función que muestra las estadísticas con números
def EstadisticasTotales(nombre):
    if os.path.exists(rutaArchivo):
        excelDocument = openpyxl.load_workbook(rutaArchivo)
        hoja = excelDocument.active
        partidas = []
        for fila in hoja.iter_rows(values_only=True):
            if nombre == nombre:   
                partidas.append(fila)
            else:
                print("\nNo se ha podido mostrar.")
        
        if partidas:
            totalPartidas = len(partidas)
            victorias = sum(1 for row in partidas if row[4] == "Ganada")
            derrotas = totalPartidas - victorias

            totalIntentos = 0
            for row in partidas:
                # Aquí se van a mostrar las estadísticas
                try:
                    intentos = int(row[3])
                    totalIntentos += intentos
                except ValueError:
                    continue    # Si no se puede convertir, pasamos

            # Calculamos la media de los intentos
            mediaIntentos = totalIntentos / totalPartidas if totalPartidas > 0 else 0

            # Calculamos el porcentaje de victorias
            porcentajeVictorias = (victorias / totalPartidas) * 100
            porcentajeDerrotas = (derrotas / totalPartidas) *100

            print(f"\nEstadísticas Totales del Juego")
            print(f"  Total de partidas: {totalPartidas}")
            print(f"  Victorias: {victorias} ({porcentajeVictorias:.2f}% de victorias)")
            print(f"  Derrotas: {derrotas} ({porcentajeDerrotas:.2f}% de derrotas)")
            print(f"  Media de intentos: {mediaIntentos:.2f}")

        else: 
            print(f"No se encontraron partidas para {nombre}.")
    else:
        print("\nNo hay estadísiticas disponibles.")


def EstadisticasFiltradasNombre(nombre):
    if os.path.exists(rutaArchivo):
        excelDocument = openpyxl.load_workbook(rutaArchivo)
        hoja = excelDocument.active
        partidas = []
        
        # Filtra las partidas en función del nombre del jugador
        for fila in hoja.iter_rows(values_only=True):
            if fila[0] == nombre:  # Nos aseguramos de que 'fila[0]' sea donde está el nombre del jugador
                partidas.append(fila)

        if partidas:
            totalPartidas = len(partidas)
            victorias = sum(1 for row in partidas if row[4] == "Ganada")  # La columna 4 es la que guarda el resultado
            derrotas = totalPartidas - victorias

            totalIntentos = 0
            for row in partidas:
                try:
                    intentos = int(row[3])  # La columna 3 es la que guarda los intentos
                    totalIntentos += intentos
                except ValueError:
                    continue  # Si no se puede convertir, pasamos

            # Calculamos la media de los intentos
            mediaIntentos = totalIntentos / totalPartidas if totalPartidas > 0 else 0

            # Calculamos el porcentaje de victorias
            porcentajeVictorias = (victorias / totalPartidas) * 100
            porcentajeDerrotas = (derrotas / totalPartidas) * 100

            print(f"\n--- Estadísiticas de {nombre} ---")
            print(f"  Total de partidas: {totalPartidas}")
            print(f"  Victorias: {victorias} ({porcentajeVictorias:.2f}% de victorias)")
            print(f"  Derrotas: {derrotas} ({porcentajeDerrotas:.2f}% de derrotas)")
            print(f"  Media de intentos: {mediaIntentos:.2f}")
        else:
            print(f"No se encontraron partidas para {nombre}.")
    else:
        print("\nNo hay estadísticas disponibles.")


def Grafico(nombre):
    if os.path.exists(rutaArchivo):
        excelDocument = openpyxl.load_workbook(rutaArchivo)
        hoja = excelDocument.active

        # Contadores para las victorias y las derrotas
        victorias = 0
        derrotas = 0

        for fila in hoja.iter_rows(values_only=True):
            if fila[0] == nombre:
                if fila[4] == "Ganada":
                    victorias += 1
                elif fila[4] == "Perdida":
                    derrotas += 1

        # Ahora verificamos si existen partidas ganadas o perdidas
        if victorias + derrotas > 0:
            # Empezamos a crear el gráfico
            plt.bar(["Ganadas", "Perdidas"], [victorias, derrotas], color=['green', 'red'])
            plt.ylabel("Nº de partidas")
            plt.title(f"Partidas Ganadas vs Perdidas de {nombre}")
            print("Cierre la ventana para continuar.")
            plt.show()
        else:
            print("No hay partidas Ganadas o Perdidas.")
    
    else:
        print("\nNo hay estadísticas disponibles.")


def Estadisticas():
    nombre = input("\n¿Cuál es tu nombre?: ").capitalize()

    while True:
        print(f"\n--- {nombre} estás en el Menú de Estadísticas ---")
        print("  1. Mostrar estadísitcas totales del juego")
        print("  2. Mostrar estadísticas filtradas por nombre")
        print(f"  3. Gráfico con partidas totales de Ganadas vs Perdidas de {nombre}")
        print("  4. Volver al Menú Principal")
        
        try: 
            opcion = int(input("\nElige una opción entre 1 y 4: "))

            if opcion == 1:
                MostrarEstadisticas() 
                EstadisticasTotales(nombre)
            elif opcion == 2:
                EstadisticasFiltradasNombre(nombre)
            elif opcion == 3:
                Grafico(nombre)
            elif opcion == 4:
                print("Saliendo del Menú de estadísiticas...")
                break
            else:
                print("Opción no válida, elige una opción entre 1 y 4")
        
        except ValueError:
            print("Error: Hay que ingresar un número que esté entre 1 y 4: ")
